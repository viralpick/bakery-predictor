"""신규 라인레벨 POS(보나비 판매 데이터_20260721.xlsx) → DAILY_COLUMNS 어댑터.

기존 `bonavi_loader`는 옛 단일시트 0520 포맷(한글 컬럼·단일 판매정보 시트·
YYYYMMDDHHMMSS)에 묶여 있다. 신규 파일은 4중 확장(4매장·전품목·컬럼·+6개월)이고
포맷도 달라(6시트·English 컬럼) 별도 어댑터를 둔다. 집계/라벨/보정 로직은
`bonavi_loader`(aggregate_daily·assign_stockout_fields·map_category 등)를 재사용한다.

게이트 A 발견 (2026-07-23, 메모리 project_new_data_ingestion_pitfall):
- 6시트 중 **판매정보2만** 판매구분(SALES_FG)/판매시간(SALES_TIME) 컬럼 순서가
  뒤바뀜(벤더 quirk). 데이터 물리 레이아웃은 6시트 동일 → **값 기준(0/1 vs 14자리)**
  으로 두 컬럼을 판별한다(헤더 순서 신뢰 금지).
- 컬럼 의미: CD_USERDEF1=할인코드, CD_USERDEF2=셋트(SS단품/ST셋트),
  SALES_FG=판매구분(0정상/1반품), SALES_TIME=판매시간(YYYYMMDDHHMMSS).
- 마스터=0526 품목정보(English): CD_ITEM/NM_ITEM(품목명)/FG_ITEM(상품구분)/
  CD_USERDEF4(당일폐기여부 Y/N).
- 재고(생산/폐기 라벨)=0526 재고정보, **2021-2025만**. 2026 H1은 sales-only.

예측 타깃(사용자 결정 2026-07-23): 당일폐기여부=Y 이면서 category != salad.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

from ..features.potential_demand import StoreHours, attach_potential_demand
from ..ingest.inventory import load_inventory
from .bonavi_loader import (
    SINGLE_FLOOR,
    _aggregate_returns,
    aggregate_daily,
    flag_bulk_lines,
    map_category,
    measure_hour_profile,
)
from .schema import DAILY_COLUMNS, validate_daily

NEW_SALES_XLSX = Path("data/internal/보나비 판매 데이터_20260721.xlsx")
MASTER_XLSX = Path("data/internal/보나비 데이터_20260526.xlsx")  # 품목정보·재고정보 시트
CLEAN_PARQUET = Path("data/internal/sales_lines_clean.parquet")
OUT_DEFAULT = Path("data/internal/bonavi_daily.parquet")
RECEIPTS_DEFAULT = Path("data/internal/bonavi_receipts.parquet")
DEFAULT_STORE_CODE = "1000000047"  # 아티제 아브뉴프랑광교점
LABEL_END = "20251231"  # 재고(폐기/생산) 라벨 가용 끝 — 이후는 sales-only

# 당일폐기=Y지만 예측 타깃 아님(구운 빵 아닌 델리). map_category salad 버킷.
EXCLUDE_CATEGORIES = frozenset({"salad"})

# String-typed 판매 라인 컬럼 (id/code/flag 계열)
_STR_COLS = frozenset(
    {"CD_PARTNER", "DT_SALE", "NO_POS", "SLIP_NO", "SLIP_LINE",
     "CD_ITEM", "CD_USERDEF1", "CD_USERDEF2", "SALES_TIME", "SALES_FG"}
)


def _fg_fraction(series: pd.Series) -> float:
    """판매구분 컬럼다움 = 값이 {'0','1'}에 속하는 비율."""
    v = series.astype("string").str.strip()
    return float(v.isin(["0", "1"]).mean())


def maybe_swap_fg_time(df: pd.DataFrame) -> pd.DataFrame:
    """헤더가 뒤바뀐 시트(판매정보2) 교정 — 값 기준으로 SALES_FG/SALES_TIME 판별.

    데이터 물리 레이아웃은 6시트 동일하나 시트2 헤더만 두 컬럼 순서가 스왑돼 있다.
    SALES_TIME 컬럼이 SALES_FG 컬럼보다 0/1다우면(=헤더가 뒤바뀐 것) 두 컬럼을 맞바꾼다.
    """
    if _fg_fraction(df["SALES_TIME"]) > _fg_fraction(df["SALES_FG"]):
        return df.rename(columns={"SALES_FG": "SALES_TIME", "SALES_TIME": "SALES_FG"})
    return df


def convert_sales_to_parquet(
    xlsx: Path | str = NEW_SALES_XLSX, out: Path | str = CLEAN_PARQUET
) -> Path:
    """신규 6시트 xlsx → 단일 클린 parquet. per-sheet 값판별 스왑 수정.

    각 시트: row0 한글 헤더(drop) / row1 English placeholder(=컬럼명) / row2+ 데이터.
    시트별로 SALES_FG가 0/1이 아니면(값 판별) SALES_TIME과 스왑한다.
    """
    xlsx, out = Path(xlsx), Path(out)
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    parts: list[pd.DataFrame] = []
    for sheet in wb.sheetnames:
        rows = wb[sheet].iter_rows(values_only=True)
        next(rows)  # 한글 헤더 drop
        names = [str(x) for x in next(rows)]  # English placeholder = 컬럼명(시트 물리순)
        df = pd.DataFrame(list(rows), columns=names)
        for col in df.columns:
            if col in _STR_COLS:
                df[col] = df[col].astype("string").str.strip()
        df = maybe_swap_fg_time(df)  # 판매정보2 헤더 스왑 교정 (값 기준)
        df["source_sheet"] = sheet
        parts.append(df)
    wb.close()
    full = pd.concat(parts, ignore_index=True)
    fg = full["SALES_FG"].astype("string").str.strip()
    if not bool(fg.isin(["0", "1"]).all()):
        raise ValueError("convert: SALES_FG에 0/1 외 값 잔존 — 스왑 판별 실패")
    out.parent.mkdir(parents=True, exist_ok=True)
    full.to_parquet(out, index=False)
    return out


def load_items_v2(master_xlsx: Path | str = MASTER_XLSX) -> pd.DataFrame:
    """0526 품목정보 → (item_id, item_name, category_id, discard_daily, is_target).

    시트는 한글 헤더(품목코드/품목명/상품구분/당일폐기여부) + English placeholder 행
    구조. placeholder 행은 상품구분="FG_ITEM"이라 SS 필터에서 자연 제거된다.
    상품구분=SS 단품만. category=map_category(품목명).
    is_target = 당일폐기여부=Y AND category not in EXCLUDE_CATEGORIES.
    """
    raw = pd.read_excel(master_xlsx, "품목정보")
    raw = raw[raw["상품구분"].astype(str).str.strip() == "SS"].copy()
    item_id = raw["품목코드"].astype("string").str.strip()
    item_name = raw["품목명"].astype("string")
    category_id = item_name.map(map_category).astype("string")
    discard_daily = raw["당일폐기여부"].astype("string").str.strip().eq("Y")
    is_target = discard_daily & ~category_id.isin(EXCLUDE_CATEGORIES)
    return pd.DataFrame(
        {
            "item_id": item_id,
            "item_name": item_name,
            "category_id": category_id,
            "discard_daily": discard_daily,
            "is_target": is_target,
        }
    ).reset_index(drop=True)


def _load_clean(store_code: str, target_items: set[str] | None) -> pd.DataFrame:
    """CLEAN_PARQUET → store + SS 필터 프레임(정상+반품 모두, 시각 파싱 포함).

    target_items가 주어지면 그 품목만 남긴다(타깃 daily 빌드용).
    반환 컬럼: store_id,date,item_id,qty,hour,minute,timestamp,receipt_id,sales_fg,discount_code
    """
    cols = ["CD_PARTNER", "DT_SALE", "NO_POS", "SLIP_NO", "CD_ITEM",
            "QT_SALE", "CD_USERDEF1", "CD_USERDEF2", "SALES_TIME", "SALES_FG"]
    df = pd.read_parquet(CLEAN_PARQUET, columns=cols)
    for c in ("CD_PARTNER", "CD_ITEM", "CD_USERDEF2", "SALES_FG", "SALES_TIME",
              "CD_USERDEF1", "DT_SALE", "NO_POS", "SLIP_NO"):
        df[c] = df[c].astype("string").str.strip()
    df = df[(df["CD_PARTNER"] == store_code) & (df["CD_USERDEF2"] == "SS")]
    if target_items is not None:
        df = df[df["CD_ITEM"].isin(target_items)]
    df = df.copy()
    df["date"] = pd.to_datetime(df["DT_SALE"], format="%Y%m%d").dt.normalize()
    ts = df["SALES_TIME"]
    df["hour"] = pd.to_numeric(ts.str.slice(8, 10), errors="coerce")
    df["minute"] = pd.to_numeric(ts.str.slice(10, 12), errors="coerce")
    df["qty"] = pd.to_numeric(df["QT_SALE"], errors="coerce").fillna(0.0).astype(float)
    df["receipt_id"] = df["DT_SALE"] + "_" + df["NO_POS"] + "_" + df["SLIP_NO"]
    return df.rename(
        columns={"CD_PARTNER": "store_id", "CD_ITEM": "item_id",
                 "SALES_FG": "sales_fg", "CD_USERDEF1": "discount_code"}
    )


def load_sales_v2(store_code: str, target_items: set[str]) -> pd.DataFrame:
    """정상매출(sales_fg=0) 단품 라인 → (store_id,date,item_id,qty,hour). bulk 제거."""
    df = _load_clean(store_code, target_items)
    df = df[df["sales_fg"] == "0"].copy()
    lines = df[["receipt_id", "item_id", "date", "qty"]]
    df = df[~flag_bulk_lines(lines).to_numpy()]
    return df[["store_id", "date", "item_id", "qty", "hour"]].reset_index(drop=True)


def load_returns_v2(store_code: str, target_items: set[str]) -> pd.DataFrame:
    """반품(sales_fg=1) 소매분 → (item_id,date,ret_qty) 집계 (대량취소 제외)."""
    df = _load_clean(store_code, target_items)
    df = df[df["sales_fg"] == "1"]
    if df.empty:
        return pd.DataFrame({"item_id": [], "date": [], "ret_qty": []})
    lines = df[["item_id", "date", "qty"]].copy()
    return _aggregate_returns(lines, single_floor=SINGLE_FLOOR)


def load_receipts_v2(store_code: str, target_items: set[str]) -> pd.DataFrame:
    """정상매출 → 시각 포함 receipts (substitution·arrival profile용). bonavi_loader 계약과 동일 컬럼."""
    df = _load_clean(store_code, target_items)
    df = df[df["sales_fg"] == "0"].dropna(subset=["hour", "minute"]).copy()
    df["hour"] = df["hour"].astype(int).clip(0, 23)
    df["minute"] = df["minute"].astype(int).clip(0, 59)
    df["timestamp"] = (
        df["date"] + pd.to_timedelta(df["hour"], unit="h")
        + pd.to_timedelta(df["minute"], unit="m")
    )
    lines = df[["receipt_id", "item_id", "date", "qty"]]
    df["is_bulk"] = flag_bulk_lines(lines).to_numpy()
    df["receipt_id"] = df["date"].dt.strftime("%Y%m%d") + "_" + df["receipt_id"]
    return df[["receipt_id", "date", "item_id", "hour", "minute", "timestamp", "qty", "is_bulk"]].assign(
        receipt_id=lambda d: d["receipt_id"].astype("string"),
        item_id=lambda d: d["item_id"].astype("string"),
    ).reset_index(drop=True)


def build_v2(
    *,
    clean_parquet: Path | str = CLEAN_PARQUET,
    master_xlsx: Path | str = MASTER_XLSX,
    store_code: str = DEFAULT_STORE_CODE,
    rename_store_id: str = "store_gw01",
    out_path: Path | str = OUT_DEFAULT,
    receipts_path: Path | str = RECEIPTS_DEFAULT,
    label_end: str = LABEL_END,
) -> Path:
    """신규 파일 → DAILY_COLUMNS daily parquet (타깃=당일폐기Y−salad, 라벨구간 <=label_end).

    convert_sales_to_parquet가 먼저 실행돼 clean_parquet이 있어야 한다.
    집계·stockout 재정의·potential_demand 보정은 bonavi_loader.aggregate_daily 재사용.
    """
    if not Path(clean_parquet).exists():
        raise FileNotFoundError(
            f"{clean_parquet} 없음 — convert_sales_to_parquet를 먼저 실행하라."
        )
    items = load_items_v2(master_xlsx)
    target_items = set(items.loc[items["is_target"], "item_id"].tolist())

    sales = load_sales_v2(store_code, target_items)
    returns = load_returns_v2(store_code, target_items)
    receipts_df = load_receipts_v2(store_code, target_items)

    # 라벨 가용 구간으로 clip (재고정보 2021-2025)
    end_ts = pd.Timestamp(pd.to_datetime(label_end, format="%Y%m%d"))
    sales = sales[sales["date"] <= end_ts].reset_index(drop=True)
    receipts_df = receipts_df[receipts_df["date"] <= end_ts].reset_index(drop=True)
    returns = (returns[returns["date"] <= end_ts].reset_index(drop=True)
               if not returns.empty else returns)

    Path(receipts_path).parent.mkdir(parents=True, exist_ok=True)
    receipts_df.to_parquet(receipts_path, index=False)

    measured_profiles = measure_hour_profile(sales)
    inventory = load_inventory(str(master_xlsx), rename_store_id)
    last_sale = (
        receipts_df.groupby(["date", "item_id"], as_index=False)["timestamp"].max()
        .rename(columns={"timestamp": "last_sale_ts"})
    )

    daily = aggregate_daily(
        sales, items[["item_id", "category_id"]], inventory, last_sale,
        returns=returns, measured_profiles=measured_profiles,
    )
    daily["store_id"] = rename_store_id
    measured_profiles = (
        {rename_store_id: next(iter(measured_profiles.values()))}
        if measured_profiles else {}
    )
    daily = daily[list(DAILY_COLUMNS.keys())]
    validate_daily(daily)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    daily.to_parquet(out_path, index=False)
    return Path(out_path)
